#!/usr/bin/env python3
"""WWMA Publishing — Finance Auto-Generator.
Baca dari PostgreSQL (single source of truth) -> regenerate workbook + P&L + neraca.
Mama (Suisui) tambah data di ledger saat: Faiz beli akun / akun deaktif / payout.
Script ini jalan tiap malam via systemd timer -> semua laporan update otomatis.

Logika keuangan DISAMAKAN dgn dashboard (`backend/app.py` `db_read_finance`) supaya
net income workbook == net income dashboard:
  net_income = payout + refund − amort − impairment − opex
  - amortisasi = hanya aset `status != 'active'`, FULL cost (bukan pro-rata)
  - refund   DIKURANG dari beban (income), bukan ditambah
  - impairment seed (upstream startswith 'upstream-') di-zero-kan (DATA-HILANG)
  - opex = 0.10
 FOREX_KEY dibaca dari env (FOREX_KEY) atau `~/.hermes-suisui/.env`, bukan hardcode.
"""
import json
import os
import sys
from datetime import date

import psycopg
from openpyxl import Workbook, load_workbook

# Rule engine bersama — satu-satunya sumber formula.
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend"))
from finance_rules import compute_finance  # noqa: E402

BASE = "/home/gamesim/shared-memory/inferhub-business/finance"
LEDGER = os.path.join(BASE, "ledger.json")
WB = os.path.join(BASE, "keuangan.xlsx")
ENV_FILE = os.path.expanduser("~/.hermes-suisui/.env")
DB_DSN = os.environ.get("UPSTREAM_DB")
if not DB_DSN:
    raise SystemExit("UPSTREAM_DB env wajib diisi (DB production). Refuse to run.")


def load_forex_key():
    """FOREX_KEY dari env; fallback baca ~/.hermes-suisui/.env. Bukan hardcode."""
    if os.environ.get("FOREX_KEY"):
        return os.environ["FOREX_KEY"].strip().strip('"').strip("'")
    try:
        with open(ENV_FILE) as f:
            for line in f:
                line = line.strip()
                if line.startswith("FOREX_KEY="):
                    return line.split("=", 1)[1].strip().strip('"').strip("'")
    except Exception:
        pass
    return None


FOREX_KEY = load_forex_key()


def load_ledger():
    """Baca seluruh ledger langsung dari PostgreSQL via UPSTREAM_DB."""
    with psycopg.connect(DB_DSN) as conn, conn.cursor() as cur:
        cur.execute("SELECT COALESCE(v, '') FROM ledger_meta WHERE k='kurs_idr_usd'")
        r = cur.fetchone()
        meta_kurs = float(r[0]) if r and r[0] else 17801.17

        cur.execute("SELECT id, upstream, qty, cost_per, curr, buy, lifespan_d, status, label, kurs_idr_usd FROM assets ORDER BY id")
        assets = [{"id": rid, "upstream": up, "qty": int(float(qty)), "cost_per": float(cost),
                   "curr": curr, "buy": (buy or "")[:10], "lifespan_d": int(float(life or 30)),
                   "status": status, "label": label or "", "kurs_idr_usd": float(kurs_a) if kurs_a is not None else None}
                  for rid, up, qty, cost, curr, buy, life, status, label, kurs_a in cur.fetchall()]

        cur.execute("SELECT id, amount_usdc, status, date FROM payouts ORDER BY date")
        payouts = [{"id": pid, "amount_usdc": float(amt or 0), "status": st or "confirmed", "date": str(d)[:10]}
                   for pid, amt, st, d in cur.fetchall()]

        cur.execute("SELECT id, upstream, qty, amount_idr, amount_usdc, label, date, kurs_idr_usd FROM refunds ORDER BY date")
        refunds = [{"id": rid, "upstream": up, "qty": int(float(qty or 0)), "amount_idr": float(aidr or 0),
                    "amount_usdc": float(ausd or 0), "label": lab or "", "date": str(d)[:10],
                    "kurs_idr_usd": float(kurs_r) if kurs_r is not None else None}
                   for rid, up, qty, aidr, ausd, lab, d, kurs_r in cur.fetchall()]

        cur.execute("SELECT id, upstream, qty, loss, label, date FROM impairments ORDER BY date")
        impairments = [{"id": iid, "upstream": up, "qty": int(float(qty or 0)), "loss": float(loss or 0),
                        "label": lab or "", "date": str(d)[:10]}
                       for iid, up, qty, loss, lab, d in cur.fetchall()]

        cur.execute("SELECT upstream_slug, count(*) AS n FROM providers WHERE status='ok' GROUP BY upstream_slug")
        providers = [{"upstream_slug": slug, "n": int(n)} for slug, n in cur.fetchall()]

    return {
        "meta": {"name": "WWMA Publishing — Ledger", "as_of": str(date.today()),
                 "kurs_idr_usd": meta_kurs, "kurs_updated": str(date.today())},
        "assets": assets, "payouts": payouts, "refunds": refunds, "impairments": impairments,
        "providers": providers,
    }


def save_ledger(L):
    with open(LEDGER, "w") as f:
        json.dump(L, f, indent=2, ensure_ascii=False)
        f.write("\n")


def fetch_live_kurs():
    """Tarik kurs IDR realtime dari forexrateapi, update ledger.json meta.
    Gagal = keep angka lama (biar report tetap jalan) tapi catat warning."""
    import urllib.request
    L = load_ledger()
    if not FOREX_KEY:
        print("⚠️ FOREX_KEY tidak ada (env / ~/.hermes-suisui/.env) — pakai kurs lama %.2f"
              % L["meta"].get("kurs_idr_usd"))
        return L["meta"]["kurs_idr_usd"]
    try:
        url = ("https://api.forexrateapi.com/v1/latest?api_key=%s"
               "&base=USD&currencies=IDR" % FOREX_KEY)
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (X11; Linux x86_64)"})
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read().decode())
        kurs = float(data["rates"]["IDR"])
        L["meta"]["kurs_idr_usd"] = kurs
        L["meta"]["kurs_updated"] = str(date.today())
        save_ledger(L)
        print("Kurs live: $1 = Rp %.2f (updated %s)" % (kurs, L["meta"]["kurs_updated"]))
    except Exception as e:
        print("⚠️ Gagal fetch kurs live (%s) — pakai lama %.2f" % (e, L["meta"].get("kurs_idr_usd")))
    return L["meta"]["kurs_idr_usd"]


def usd_eq(amount, curr, kurs):
    return amount if curr == "USD" else amount / kurs


def main():
    L = load_ledger()
    kurs = fetch_live_kurs()          # selalu tarik kurs realtime, update meta
    today = L["meta"]["as_of"]

    res = compute_finance(
        assets=L["assets"], payouts=L["payouts"], refunds=L["refunds"],
        impairments=L["impairments"], kurs_meta=kurs, providers=L["providers"],
    )
    total_payout = res["total_payout"]
    n_payout = res["n_payout"]
    total_amort_usd = res["amort_usd"]
    amort_assets = res["amort_assets"]
    total_imp_loss_usd = res["total_imp_loss_usd"]
    impaired_rows = res["impairments"]
    total_refund_usd = res["total_refund_usd"]
    refund_rows = res["refunds"]
    opex = res["opex"]
    net_income = res["net_income"]
    asset_list = res["assets"]
    total_capital_usd = res["total_capital_usd"]
    total_akun = res["total_asset_qty"]
    total_akun_aktif = res["aktif"]

    # ---- Simpan ringkasan ke JSON (untuk ref script lain / debug) ----
    summary = {
        "as_of": today,
        "kurs": kurs,
        "total_payout_usd": round(total_payout, 2),
        "n_payout": n_payout,
        "total_refund_usd": total_refund_usd,
        "total_amort_usd": total_amort_usd,
        "amort_assets": [{**a, "cost_usd": a["cost_usd"]} for a in amort_assets],
        "total_imp_loss_usd": total_imp_loss_usd,
        "opex": opex,
        "net_income": net_income,
        "total_capital_usd": round(total_capital_usd, 2),
        "total_akun_assets": total_akun,
        "total_akun_aktif": total_akun_aktif,
    }
    print(json.dumps(summary, indent=2))

    # ---- Tulis workbook ----
    wb = Workbook()

    # Income Statement (USD only)
    ws = wb.active
    ws.title = "Income Statement"
    rows_inc = [
        ["LAPORAN LABA RUGI", None],
        [f"Periode: 05 Jul – {today} • USD ONLY • Kurs $1 = Rp {kurs:,.2f}", None],
        [None, None],
        ["PENDAPATAN", None],
        [f"Revenue – Earnings USDC ({n_payout} payout)", total_payout],
        [f"Refund ({len(refund_rows)}) — pengurang beban", total_refund_usd],
        ["Total Pendapatan", round(total_payout + total_refund_usd, 2)],
        [None, None],
        ["BEBAN POKOK (COGS)", None],
        ["Amortisasi aset (non-active, full cost)", total_amort_usd],
        ["Total COGS (USD)", total_amort_usd],
        [None, None],
        ["LABA KOTOR", round(total_payout + total_refund_usd - total_amort_usd, 2)],
        [None, None],
        ["BEBAN OPERASIONAL", None],
        ["Beban Bank/Gas Fee (est.)", opex],
        [f"Impairment {len(impaired_rows)} akun invalid ({round(total_imp_loss_usd, 2)} USD)",
         total_imp_loss_usd],
        ["Total Beban Operasional", round(opex + total_imp_loss_usd, 2)],
        [None, None],
        ["LABA BERSIH (NET INCOME)", net_income],
    ]
    for r in rows_inc:
        ws.append(r)

    # Asset Register (USD eq — full USD sesuai operator)
    wsa = wb.create_sheet("Asset Register")
    wsa.append(["ID", "Upstream", "Qty", "Cost/unit (native)", "Cost/unit (USD)", "Total USD eq", "Buy", "Lifespan_d", "Status"])
    for a in asset_list:
        cost_usd = usd_eq(a["cost_per"], a["curr"], kurs)
        wsa.append([a["id"], a["upstream"], a["qty"], a["cost_per"], round(cost_usd, 4),
                    round(a["cost_usd"], 4), a["buy"], a["lifespan_d"], a["status"]])
    wsa.append(["TOTAL", "", total_akun, "", "", round(total_capital_usd, 2), "", "", ""])

    # Payout register
    wsp = wb.create_sheet("Payouts")
    wsp.append(["Tanggal", "USD", "Status", "Note"])
    for p in L["payouts"]:
        wsp.append([p["date"], p.get("amount_usdc", p.get("usd", 0)), p.get("status", ""), p.get("note", "")])
    wsp.append(["TOTAL", total_payout, "", ""])

    # Refund register
    wsr = wb.create_sheet("Refunds")
    wsr.append(["ID", "Upstream", "Qty", "IDR", "USD", "Refund USD", "Label", "Date"])
    for rd in refund_rows:
        wsr.append([rd["id"], rd["upstream"], rd["qty"], rd["amount_idr"], rd["amount_usdc"],
                    rd["refund_usd"], rd["label"], rd["date"]])
    wsr.append(["TOTAL", "", "", "", "", total_refund_usd, "", ""])

    # Impairment register
    wsi = wb.create_sheet("Impairments")
    wsi.append(["ID", "Label", "Qty", "Loss", "Loss USD", "Curr", "Date"])
    for im in impaired_rows:
        wsi.append([im["id"], im["label"], im["qty"], im["loss"], im["loss_usd"],
                    im.get("curr", "IDR"), im["date"]])
    wsi.append(["TOTAL", "", "", total_imp_loss_usd, "", "", ""])

    wb.save(WB)
    print("Workbook regenerated:", WB)


if __name__ == "__main__":
    main()